import json
import os
from datetime import date
from datetime import datetime
today = date.today()
datetoday = today.strftime("%Y-%m-%d")
















###############
'''reset staging'''
import openpyxl as xl;
filename ="_Staging/Staging.xlsx"
wb= xl.load_workbook(filename)

listsheets=wb.sheetnames
print(wb.sheetnames)
ws2 = wb.active
print(ws2)
# workbook.remove(operations_sheet)

for i in listsheets:
 if i !="Sheet1":  
  ws2 = wb[i]
  print(ws2)
  wb.remove(ws2)
  print(wb.sheetnames)
  wb.save(str(filename))
 else:
    pass