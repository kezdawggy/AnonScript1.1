#####READ the GL-np file - Gather UW Year

import openpyxl as xl;
filename ="_Staging/Staging.xlsx"
wb = xl.load_workbook(filename)
print(wb.sheetnames)
GLsheet = wb["GL-np"]

year=[]


for value in GLsheet.iter_rows(min_row=5, 
                          max_row=17,
                          min_col=1,
                          max_col=1,
                          values_only=True
                          
                                ):

       year.append(value[0])
header0=year.pop(0)
print([header0])
print(year)
#     years={}
#     years=
#     data={}
#     print(value)