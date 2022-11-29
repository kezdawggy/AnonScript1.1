#####READ the MA-np file - Gather Statistical Data
import openpyxl as xl;
filename ="_Staging/Staging.xlsx"
wb = xl.load_workbook(filename)
print(wb.sheetnames)
MAsheet = wb["MA-np"]

prem_stats=[]

# print(trianglesheet)
for value in MAsheet.iter_rows(min_row=5, 
                          max_row=17,
                          min_col=2,
                          max_col=14,
                          values_only=True
                          
                                ):

       prem_stats.append(list(value))
headerstats=list(prem_stats.pop(0))
print(headerstats)
print(prem_stats)
#     years={}
#     years=
#     data={}
#     print(value)