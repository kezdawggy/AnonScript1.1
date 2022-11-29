import openpyxl as xl;

#####READ the MA-np file - Gather UW Year

filename ="_Staging/Staging.xlsx"
wb = xl.load_workbook(filename)
print(wb.sheetnames)
MAsheet = wb["MA-np"]

year=[]

# print(trianglesheet)
for value in MAsheet.iter_rows(min_row=5, 
                          max_row=17,
                          min_col=1,
                          max_col=1,
                          values_only=True
                          
                                ):

       year.append(value[0])
header0=year.pop(0)
print([header0])
print(year)
#     years={}
#     years=
#     data={}
#     print(value)