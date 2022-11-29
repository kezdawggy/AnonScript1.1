###capture the source triangle data
import openpyxl as xl;
  
# opening the source excel file
filename ="_Staging/Staging.xlsx"
wb = xl.load_workbook(filename)
print(wb.sheetnames)
trianglesheet = wb['Triangle Data 2020']

print(trianglesheet)
for value in trianglesheet.iter_rows(min_row=2, 
                          max_row=2,
                          min_col=4,
                          max_col=15,
                          values_only=True):
    print(value)
    
TriangleMonth=list(value)   
print(TriangleMonth)