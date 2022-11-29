###load up staging excel.

# importing openpyxl module
import openpyxl as xl;
  
# opening the source excel file
filename ="_Inputs/Howden_CompanyXYZ_2021_Data_orig.xlsx"
wb1 = xl.load_workbook(filename)

sourcelistsheets=wb1.sheetnames
print(wb1.sheetnames)

for i in sourcelistsheets:

    if i =="GL-np": 

        # ws1 = wb1.worksheets["GL-np"]
        ws1 = wb1[i]
        # opening the destination excel file 
        filename1 ="_Staging/Staging.xlsx"
        wb2 = xl.load_workbook(filename1)
        ws2 = wb2.active

        ws2 = wb2.create_sheet(title=i)
        ws2 = wb2[i]  
        # calculate total number of rows and 
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column
  
        # copying the cell values from source 
        # excel file to destination excel file
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
            # reading cell value from source excel file
                c = ws1.cell(row = i, column = j)
  
                # writing the read value to destination excel file
                ws2.cell(row = i, column = j).value = c.value
  
            # saving the destination excel file
        wb2.save(str(filename1))
        
    elif i =="MA-np": 

        # ws1 = wb1.worksheets["GL-np"]
        ws1 = wb1[i]
        # opening the destination excel file 
        filename1 ="_Staging/Staging.xlsx"
        wb2 = xl.load_workbook(filename1)
        ws2 = wb2.active

        ws2 = wb2.create_sheet(title=i)
        ws2 = wb2[i]  
        # calculate total number of rows and 
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column
  
        # copying the cell values from source 
        # excel file to destination excel file
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
            # reading cell value from source excel file
                c = ws1.cell(row = i, column = j)
  
                # writing the read value to destination excel file
                ws2.cell(row = i, column = j).value = c.value
  
            # saving the destination excel file
        wb2.save(str(filename1))
        
    elif i =="Triangle Data 2020": 

        # ws1 = wb1.worksheets["GL-np"]
        ws1 = wb1[i]
        # opening the destination excel file 
        filename1 ="_Staging/Staging.xlsx"
        wb2 = xl.load_workbook(filename1)
        ws2 = wb2.active

        ws2 = wb2.create_sheet(title=i)
        ws2 = wb2[i]  
        # calculate total number of rows and 
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column
  
        # copying the cell values from source 
        # excel file to destination excel file
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
            # reading cell value from source excel file
                c = ws1.cell(row = i, column = j)
  
                # writing the read value to destination excel file
                ws2.cell(row = i, column = j).value = c.value
  
            # saving the destination excel file
        wb2.save(str(filename1))      
        
    elif i =="Sheet1":  
      ws2 = wb2[i]
      print(wb.sheetnames)
      wb2.remove(ws2)    
      wb2.save(str(filename1)) 