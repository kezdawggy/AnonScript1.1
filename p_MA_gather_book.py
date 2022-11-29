import openpyxl as xl;

#####READ the MA-np file - Gatherbook Data

filename ="_Staging/Staging.xlsx"
wb = xl.load_workbook(filename)
print(wb.sheetnames)
GLsheet = wb["MA-np"]

booked_data=[]


for value in GLsheet.iter_rows(min_row=5, 
                          max_row=17,
                          min_col=16,
                          max_col=19,
                          values_only=True
                          
                                ):

       booked_data.append(list(value))
bookedheader=list(booked_data.pop(0))
print(bookedheader)
print(booked_data)
#     years={}
#     years=
#     data={}
#     print(value)